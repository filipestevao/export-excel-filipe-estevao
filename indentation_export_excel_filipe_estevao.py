# ==============================================================================
# Script Name: Export Excel by Filipe Estevao
# Description: Python script for data export to Excel.
# 
# Copyright (c) 2026 Filipe Estevão
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
# ==============================================================================

__title__ = "Export Excel by Filipe Estevao"
__version__ = "1.0.4"
__author__ = "Filipe Estevao"
__status__ = "Production"
__url__ = "https://github.com/filipestevao/export-excel-filipe-estevao"

import json
import math
import os
import subprocess
import sys
import urllib.request

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.data_source import (
    AxDataSource,
    NumData,
    NumDataSource,
    NumRef,
    NumVal,
    StrData,
    StrRef,
    StrVal,
)
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter, quote_sheetname

from antonpaar import connect_to_indentation
from antonpaar.script_tools import info


CURVE_TYPES = ['ctCombinated', 'ctStatic', 'ctDataCurves', 'ctFreeCurves']
STAT_ROWS = ['Mean', 'Std dev']
# To show additional statistic rows, add any of these names to STAT_ROWS:
# 'Min', 'Max', 'N'
X_NAMES = ['Pd', 'Penetration depth']
Y_NAMES = ['Fn', 'Normal force']
SUMMARY_CHART_NAMES = ('HIT', 'EIT', 'E*')
GROUP_COLORS = (
    '2B78C4', 'B0689C', '53A6A6', 'FF9DA7', 'FF9B00',
    '71C840', 'E15759', '9C755F', '8C8C8C', '28292E')


def is_number(value):
    return isinstance(value, (int, float)) and not math.isnan(value)


def stats(values):
    values = [v for v in values if is_number(v)]
    if not values:
        return {'Mean': None, 'Std dev': None, 'Min': None, 'Max': None, 'N': 0}
    mean = sum(values) / len(values)
    if len(values) > 1:
        variance = sum((v - mean) ** 2 for v in values) / (len(values) - 1)
        std_dev = math.sqrt(variance)
    else:
        std_dev = 0
    return {
        'Mean': mean,
        'Std dev': std_dev,
        'Min': min(values),
        'Max': max(values),
        'N': len(values),
    }


def curve_dimension(curves, curve_type, names):
    if curve_type not in curves:
        return None
    names = [n.lower() for n in names]
    curve_items = sorted(
        curves[curve_type].items(),
        key=lambda item: int(item[0]),
    )
    for _, meta in curve_items:
        display = meta.get('DisplayName', '').lower()
        short = meta.get('ShortName', '').lower()
        if display in names or short in names:
            return int(meta.get('DimIndex'))
    for _, meta in curve_items:
        display = meta.get('DisplayName', '').lower()
        short = meta.get('ShortName', '').lower()
        if any(n in display or n in short for n in names):
            return int(meta.get('DimIndex'))
    return None


def curve_unit(curves, curve_type, dim_index):
    meta = curves[curve_type].get(str(dim_index), {})
    unit = meta.get('PhysicalUnit', {})
    return unit.get('Symbol', ''), unit.get('SICoef', 1.0) or 1.0


def curve_name(curves, curve_type, dim_index):
    meta = curves[curve_type].get(str(dim_index), {})
    return (
        meta.get('ShortName')
        or meta.get('DisplayName')
        or 'Dim %d' % dim_index
    )


def selected_acquisitions(groups):
    selected = []
    for group_id in groups['indexes']:
        group = groups['groups'][group_id]
        acquisitions = []
        for acquisition_index, data_id in enumerate(group['indexes'], 1):
            acquisition = group['data'][data_id]
            if acquisition.get('relevant'):
                acquisitions.append((data_id, acquisition, acquisition_index))
        if acquisitions:
            selected.append((group_id, group, acquisitions))
    return selected


def get_curve_data(server, doc_id, data_id, curve_type):
    first = server.curves.getdata(
        doc_id=doc_id,
        data_id=data_id,
        page_index=0,
        page_size=1,
        curve_type=curve_type)
    count = first.get('count', 0)
    if count == 0:
        return []
    data = server.curves.getdata(
        doc_id=doc_id,
        data_id=data_id,
        page_index=0,
        page_size=count,
        curve_type=curve_type)
    return data.get('data', [])


def get_breakpoints(server, doc_id, data_id, count):
    try:
        breakpoints = server.curves.breakpoints(
            doc_id=doc_id,
            data_id=data_id,
            curviline=True).get('breakpoints', [])
    except Exception:
        breakpoints = []
    breakpoints = [int(i) for i in breakpoints if 0 <= int(i) < count]
    if 0 not in breakpoints:
        breakpoints.insert(0, 0)
    if count and count - 1 not in breakpoints:
        breakpoints.append(count - 1)
    return sorted(set(breakpoints))


def measurement_name(group, acquisition, acquisition_index):
    name = acquisition.get('name') or 'Measurement %d' % acquisition_index
    return '%s - %s' % (group['name'], name)


def curve_measurement_name(group, acquisition, acquisition_index):
    name = acquisition.get('name') or str(acquisition_index)
    return '%s - %s' % (group['name'], name)


def summary_chart_name(parameter):
    compact = parameter.upper().replace(' ', '')
    if 'HIT' in compact:
        return 'HIT'
    if 'EIT' in compact:
        return 'EIT'
    if 'E*' in compact:
        return 'E*'
    return None


def color_for_group(group_name, group_colors):
    if group_name not in group_colors:
        color_index = len(group_colors)
        group_colors[group_name] = varied_color(
            GROUP_COLORS[color_index % len(GROUP_COLORS)],
            color_index // len(GROUP_COLORS))
    return group_colors[group_name]


def varied_color(hex_color, variation):
    if variation == 0:
        return hex_color

    values = [int(hex_color[i:i + 2], 16) for i in range(0, 6, 2)]
    if variation % 2:
        factor = max(0.35, 1.0 - 0.15 * ((variation + 1) // 2))
        values = [int(value * factor) for value in values]
    else:
        factor = min(0.55, 0.15 * (variation // 2))
        values = [int(value + (255 - value) * factor) for value in values]
    return ''.join('%02X' % value for value in values)


def parameter_value(
    server,
    doc_id,
    data_id,
    param_id,
    unit_factor,
    cycle_index=None,
    cycle_key='cycle_index',
):
    kwargs = {
        'doc_id': doc_id,
        'data_id': data_id,
        'param_id': param_id,
    }
    if cycle_index is not None:
        kwargs[cycle_key] = cycle_index
    value = server.parameters.getvalue(**kwargs)
    if isinstance(value, dict):
        if value.get('defined', True) is False:
            return None
        value = value.get('value')
    if is_number(value):
        return value / unit_factor
    return None


def safe_acquisition_analyses(server, doc_id, data_id):
    try:
        return server.acquisitions.analyses(
            doc_id=doc_id,
            acquisition_id=data_id).get('result', [])
    except Exception as error:
        info(' - skipping analyses for acquisition %s: %s' % (data_id, error))
        return []


def acquisition_result_value(
    server,
    doc_id,
    acquisition_id,
    analysis_id,
    param_id,
    unit_factor,
):
    for data_id in (analysis_id, acquisition_id):
        try:
            value = parameter_value(
                server,
                doc_id,
                data_id,
                param_id,
                unit_factor,
            )
        except Exception:
            continue
        if value is not None:
            return value
    return None




def progress_average_segment(segment_x, segment_y, target_count):
    progress = np.linspace(0.0, 1.0, target_count)
    x_grid = []
    y_grid = []
    for x, y in zip(segment_x, segment_y):
        source_progress = np.linspace(0.0, 1.0, len(x))
        x_grid.append(np.interp(progress, source_progress, x))
        y_grid.append(np.interp(progress, source_progress, y))
    return (
        np.maximum(np.mean(np.array(x_grid), axis=0), 0),
        np.maximum(np.mean(np.array(y_grid), axis=0), 0),
    )


def average_segment(segment_x, segment_y, target_count):
    return progress_average_segment(segment_x, segment_y, target_count)


def write_results_sheet(wb, server, doc_id, selected, analyses_classes):
    ws = wb.create_sheet('Results')
    ws.freeze_panes = 'B1'
    header_fill = PatternFill('solid', fgColor='FCE4D6')
    section_fill = PatternFill('solid', fgColor='D9EAF7')
    summary_chart_data = {
        name: {'unit': '', 'groups': []}
        for name in SUMMARY_CHART_NAMES
    }
    row = 1

    ws.cell(row, 1, 'Indentation results')
    ws.cell(row, 1).font = Font(bold=True)
    row += 2

    for _, group, acquisitions in selected:
        ws.cell(row, 1, group['name'])
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).fill = section_fill
        row += 1

        columns = []
        columns_by_key = {}
        acquisition_names = []
        for data_id, acquisition, acquisition_index in acquisitions:
            acquisition_name = measurement_name(
                group,
                acquisition,
                acquisition_index,
            )
            acquisition_names.append(acquisition_name)
            analyses = safe_acquisition_analyses(server, doc_id, data_id)
            for analysis in analyses:
                analysis_class = analyses_classes.get(analysis['class_id'])
                if not analysis_class:
                    continue
                for parameter in analysis_class.get('parameters', []):
                    unit_factor = parameter.get('unit_factor', 1.0) or 1.0
                    value = acquisition_result_value(
                        server,
                        doc_id,
                        data_id,
                        analysis['id'],
                        parameter['id'],
                        unit_factor)
                    if value is not None:
                        key = (
                            analysis_class.get('class_name', ''),
                            parameter.get('name', ''),
                            parameter.get('unit', ''),
                        )
                        if key not in columns_by_key:
                            columns_by_key[key] = {
                                'analysis': analysis_class.get(
                                    'class_name',
                                    '',
                                ),
                                'parameter': parameter.get('name', ''),
                                'unit': parameter.get('unit', ''),
                                'values': [],
                                'acquisitions': {},
                            }
                            columns.append(columns_by_key[key])
                        columns_by_key[key]['values'].append(value)
                        columns_by_key[key]['acquisitions'][
                            acquisition_name
                        ] = value

        for column in columns:
            column['stats'] = stats(column['values'])

        group_chart_names = set()
        for column in columns:
            chart_name = summary_chart_name(column['parameter'])
            if not chart_name or chart_name in group_chart_names:
                continue
            group_chart_names.add(chart_name)
            if not summary_chart_data[chart_name]['unit']:
                summary_chart_data[chart_name]['unit'] = column['unit']
            summary_chart_data[chart_name]['groups'].append({
                'group': group['name'],
                'mean': column['stats']['Mean'],
                'std_dev': column['stats']['Std dev'],
            })

        if not columns:
            ws.cell(
                row,
                1,
                'No analysis results found for selected measurements',
            )
            row += 2
            continue

        ws.cell(row, 1, '')
        ws.cell(row + 1, 1, '')
        for col, column in enumerate(columns, 2):
            title = column['parameter']
            if column['unit']:
                title += ' [%s]' % column['unit']
            ws.cell(row, col, title)
            ws.cell(row + 1, col, column['analysis'])
            for r in range(row, row + 2):
                ws.cell(r, col).fill = header_fill
                ws.cell(r, col).font = Font(bold=True)

        for offset, stat_name in enumerate(STAT_ROWS, 2):
            ws.cell(row + offset, 1, stat_name)
            ws.cell(row + offset, 1).font = Font(bold=True)
            for col, column in enumerate(columns, 2):
                ws.cell(row + offset, col, column['stats'][stat_name])

        row += len(STAT_ROWS) + 2
        for acquisition_name in acquisition_names:
            ws.cell(row, 1, acquisition_name)
            for col, column in enumerate(columns, 2):
                ws.cell(row, col, column['acquisitions'].get(acquisition_name))
            row += 1

        row += 2

    # Auto-size first column A to its content. openpyxl can't perfectly
    # match Excel AutoFit, so compute measured width and apply padding.
    # Build list of longest line lengths for each non-empty cell.
    lengths = []
    for cell in ws['A']:
        if cell.value is None:
            continue
        text = str(cell.value)
        # splitlines() returns empty list for empty string; fall back to
        # original text to ensure there's at least one item.
        lines = text.splitlines() or [text]
        lengths.append(max(len(line) for line in lines))
    max_len = max(lengths) if lengths else 16
    width = math.ceil(max_len * 1.05) + 2
    width = min(max(width, 8), 255)
    col_dim = ws.column_dimensions[get_column_letter(1)]
    col_dim.width = width

    for col in range(2, min(ws.max_column, 24) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    return summary_chart_data


def choose_curve(curves):
    for curve_type in CURVE_TYPES:
        x_dim = curve_dimension(curves, curve_type, X_NAMES)
        y_dim = curve_dimension(curves, curve_type, Y_NAMES)
        if x_dim is not None and y_dim is not None:
            return curve_type, x_dim, y_dim
    raise RuntimeError(
        'Could not find Pd and Fn curve dimensions in the document'
    )


def write_curves_sheet(
    wb,
    server,
    doc_id,
    selected,
    curves,
    curve_type,
    x_dim,
    y_dim,
    group_colors,
):
    ws = wb.create_sheet('Curves')
    ws.freeze_panes = 'A3'
    x_unit, x_factor = curve_unit(curves, curve_type, x_dim)
    y_unit, y_factor = curve_unit(curves, curve_type, y_dim)
    x_name = curve_name(curves, curve_type, x_dim)
    y_name = curve_name(curves, curve_type, y_dim)
    exported = []
    chart = None
    col = 1

    for _, group, acquisitions in selected:
        for data_id, acquisition, acquisition_index in acquisitions:
            rows = get_curve_data(server, doc_id, data_id, curve_type)
            if not rows:
                continue
            breakpoints = get_breakpoints(server, doc_id, data_id, len(rows))
            display_name = curve_measurement_name(
                group,
                acquisition,
                acquisition_index,
            )
            ws.cell(1, col, display_name)
            ws.cell(2, col, '%s [%s]' % (x_name, x_unit) if x_unit else x_name)
            ws.cell(
                2,
                col + 1,
                '%s [%s]' % (y_name, y_unit) if y_unit else y_name,
            )
            x_values = []
            y_values = []
            for row_index, point in enumerate(rows, 3):
                x_value = point[x_dim] / x_factor
                y_value = point[y_dim] / y_factor
                ws.cell(row_index, col, x_value)
                ws.cell(row_index, col + 1, y_value)
                if is_number(x_value) and is_number(y_value):
                    x_values.append(x_value)
                    y_values.append(y_value)
            exported.append({
                'name': display_name,
                'group': group['name'],
                'col': col,
                'count': len(rows),
                'breakpoints': breakpoints,
                'x': np.array(x_values, dtype=float),
                'y': np.array(y_values, dtype=float),
            })
            col += 3

    if exported:
        chart = ScatterChart()
        chart.scatterStyle = 'lineMarker'
        chart.title = 'Indentation curves'
        chart.x_axis.title = '%s [%s]' % (x_name, x_unit) if x_unit else x_name
        chart.y_axis.title = '%s [%s]' % (y_name, y_unit) if y_unit else y_name
        chart.x_axis.scaling.min = 0
        chart.y_axis.scaling.min = 0
        chart.width = 24
        chart.height = 14
        for item in exported:
            x_values = Reference(
                ws,
                min_col=item['col'],
                min_row=3,
                max_row=item['count'] + 2,
            )
            y_values = Reference(
                ws,
                min_col=item['col'] + 1,
                min_row=3,
                max_row=item['count'] + 2,
            )
            series = Series(y_values, x_values, title=item['name'])
            color = color_for_group(item['group'], group_colors)
            series.graphicalProperties.line.solidFill = color
            series.marker.graphicalProperties.solidFill = color
            series.marker.graphicalProperties.line.solidFill = color
            cache_num_ref(series.xVal, item['x'].tolist())
            cache_num_ref(series.yVal, item['y'].tolist())
            chart.series.append(series)

    return exported, x_name, x_unit, y_name, y_unit, chart


def write_average_curves_sheet(
    wb,
    exported,
    x_name,
    x_unit,
    y_name,
    y_unit,
    group_colors,
):
    ws = wb.create_sheet('Average curves')
    groups = []
    for item in exported:
        if len(item['x']) < 2 or len(item['y']) < 2:
            continue
        if item['group'] not in groups:
            groups.append(item['group'])

    if not groups:
        ws.cell(2, 1, 'No curve data available for averaging')
        return None

    chart = ScatterChart()
    chart.scatterStyle = 'lineMarker'
    chart.title = 'Average indentation curves'
    chart.x_axis.title = '%s [%s]' % (x_name, x_unit) if x_unit else x_name
    chart.y_axis.title = '%s [%s]' % (y_name, y_unit) if y_unit else y_name
    chart.width = 24
    chart.height = 14
    col = 1

    for group_name in groups:
        usable_items = [
            item
            for item in exported
            if (
                item['group'] == group_name
                and len(item['x']) >= 2
                and len(item['y']) >= 2
            )
        ]
        if not usable_items:
            continue

        ws.cell(1, col, group_name)
        ws.cell(
            2,
            col,
            'Mean %s [%s]' % (x_name, x_unit)
            if x_unit
            else 'Mean %s' % x_name,
        )
        ws.cell(
            2,
            col + 1,
            'Mean %s [%s]' % (y_name, y_unit)
            if y_unit
            else 'Mean %s' % y_name,
        )

        averaged_x = []
        averaged_y = []
        segment_count = max(
            len(item['breakpoints']) - 1
            for item in usable_items
        )
        for segment_index in range(segment_count):
            segment_x = []
            segment_y = []
            target_count = 0
            for item in usable_items:
                if segment_index + 1 >= len(item['breakpoints']):
                    continue
                start = item['breakpoints'][segment_index]
                stop = item['breakpoints'][segment_index + 1]
                x = np.maximum(item['x'][start:stop + 1], 0)
                y = np.maximum(item['y'][start:stop + 1], 0)
                if segment_index == 0:
                    positive = np.where((x > 0) | (y > 0))[0]
                    if len(positive):
                        x = x[positive[0]:]
                        y = y[positive[0]:]
                    x = np.insert(x, 0, 0.0)
                    y = np.insert(y, 0, 0.0)
                if len(x) < 2 or len(y) < 2:
                    continue
                segment_x.append(x)
                segment_y.append(y)
                target_count = max(target_count, len(x))

            if target_count < 2 or not segment_x:
                continue

            x_mean, y_mean = average_segment(segment_x, segment_y, target_count)
            if averaged_x:
                x_mean = x_mean[1:]
                y_mean = y_mean[1:]
            averaged_x.extend(x_mean.tolist())
            averaged_y.extend(y_mean.tolist())

        if not averaged_x:
            continue

        for index, x_value in enumerate(averaged_x, 3):
            ws.cell(index, col, float(x_value))
            ws.cell(index, col + 1, float(averaged_y[index - 3]))

        last_row = len(averaged_x) + 2
        x_values = Reference(ws, min_col=col, min_row=3, max_row=last_row)
        y_values = Reference(ws, min_col=col + 1, min_row=3, max_row=last_row)
        series = Series(y_values, x_values, title=group_name)
        color = color_for_group(group_name, group_colors)
        series.graphicalProperties.line.solidFill = color
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color
        cache_num_ref(series.xVal, averaged_x)
        cache_num_ref(series.yVal, averaged_y)
        chart.series.append(series)
        col += 3

    if chart.series:
        return chart

    return None


def chart_title(name, unit):
    return '%s [%s]' % (name, unit) if unit else name


def range_ref(ws, column, first_row, last_row):
    return '%s!$%s$%d:$%s$%d' % (
        quote_sheetname(ws.title),
        get_column_letter(column),
        first_row,
        get_column_letter(column),
        last_row)


def number_cache(values):
    points = [
        NumVal(idx=index, v=value)
        for index, value in enumerate(values)
        if is_number(value)
    ]
    return NumData(ptCount=len(values), pt=points)


def string_cache(values):
    points = [
        StrVal(idx=index, v='' if value is None else str(value))
        for index, value in enumerate(values)
    ]
    return StrData(ptCount=len(values), pt=points)


def cache_num_ref(data_source, values):
    if data_source and data_source.numRef:
        data_source.numRef.numCache = number_cache(values)


def set_chart_axis_ids(chart, base_id):
    x_axis_id = base_id + 1
    y_axis_id = base_id + 2
    chart.x_axis.axId = x_axis_id
    chart.y_axis.axId = y_axis_id
    chart.x_axis.crossAx = y_axis_id
    chart.y_axis.crossAx = x_axis_id
    chart.x_axis.axPos = 'b'
    chart.y_axis.axPos = 'l'
    for axis in (chart.x_axis, chart.y_axis):
        axis.delete = False
        axis.tickLblPos = 'nextTo'
        axis.crosses = 'autoZero'
        axis.numFmt = 'General'
    if isinstance(chart, ScatterChart):
        chart.x_axis.crossBetween = 'midCat'
        chart.y_axis.crossBetween = 'midCat'
    else:
        chart.y_axis.crossBetween = 'between'


def style_summary_bar_chart(chart):
    chart.y_axis.scaling.min = 0

    # Set overall chart border
    if chart.plot_area.spPr is None:
        chart.plot_area.spPr = GraphicalProperties()
    if chart.plot_area.spPr.ln is None:
        chart.plot_area.spPr.ln = LineProperties()
    chart.plot_area.spPr.ln.solidFill = '000000'

    chart.y_axis.majorGridlines.spPr = GraphicalProperties()
    chart.y_axis.majorGridlines.spPr.line.solidFill = 'C5C5C5'


def add_summary_bar_chart(
    ws,
    chart_name,
    chart_info,
    anchor,
    data_col,
    axis_base_id,
):
    groups = chart_info['groups']
    unit = chart_info['unit']

    ws.cell(2, data_col, chart_name)
    ws.cell(3, data_col, 'Group')
    ws.cell(3, data_col + 1, 'Mean')
    ws.cell(3, data_col + 2, 'Std dev')

    for row_offset, item in enumerate(groups, 4):
        ws.cell(row_offset, data_col, item['group'])
        ws.cell(row_offset, data_col + 1, item['mean'])
        ws.cell(row_offset, data_col + 2, item['std_dev'])

    if not groups:
        ws.cell(4, data_col, 'No %s data found' % chart_name)
        return

    first_row = 4
    last_row = first_row + len(groups) - 1
    chart = BarChart()
    chart.title = chart_title(chart_name, unit)
    chart.y_axis.title = chart_title(chart_name, unit)
    chart.legend = None
    chart.width = 12
    chart.height = 8
    set_chart_axis_ids(chart, axis_base_id)
    style_summary_bar_chart(chart)
    chart.type = 'col'
    chart.grouping = 'clustered'
    chart.add_data(
        Reference(ws, min_col=data_col + 1, min_row=3, max_row=last_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=data_col,
            min_row=first_row,
            max_row=last_row,
        )
    )
    chart.series[0].cat = AxDataSource(strRef=StrRef(
        f=range_ref(ws, data_col, first_row, last_row),
        strCache=string_cache([item['group'] for item in groups])))
    cache_num_ref(chart.series[0].val, [item['mean'] for item in groups])

    std_values = [item['std_dev'] for item in groups]
    std_ref = range_ref(ws, data_col + 2, first_row, last_row)
    chart.series[0].errBars = ErrorBars(
        errDir='y',
        errBarType='both',
        errValType='cust',
        plus=NumDataSource(
            numRef=NumRef(f=std_ref, numCache=number_cache(std_values))
        ),
        minus=NumDataSource(
            numRef=NumRef(f=std_ref, numCache=number_cache(std_values))
        ),
    )
    ws.add_chart(chart, anchor)


def write_charts_sheet(
    ws,
    doc_name,
    summary_chart_data,
    curve_chart,
    average_curve_chart,
):
    ws.cell(1, 1, doc_name)
    ws.cell(1, 1).font = Font(bold=True)

    anchors = {'HIT': 'A3', 'EIT': 'I3', 'E*': 'Q3'}
    data_col = 30
    axis_base_id = 1000
    for chart_name in SUMMARY_CHART_NAMES:
        add_summary_bar_chart(
            ws,
            chart_name,
            summary_chart_data[chart_name],
            anchors[chart_name],
            data_col,
            axis_base_id)
        data_col += 4
        axis_base_id += 100

    if curve_chart is not None:
        set_chart_axis_ids(curve_chart, axis_base_id)
        ws.add_chart(curve_chart, 'A20')
        axis_base_id += 100
    if average_curve_chart is not None:
        set_chart_axis_ids(average_curve_chart, axis_base_id)
        ws.add_chart(average_curve_chart, 'A48')


def export_selected_indentation_excel(server, doc_id):
    docs = server.docs()
    doc = docs['docs'][doc_id]
    doc_path = doc.get('path') or doc.get('name') or 'indentation_export'
    export_path = os.path.splitext(doc_path)[0] + '.xlsx'
    groups = server.groups(doc_id=doc_id)
    selected = selected_acquisitions(groups)
    if not selected:
        raise RuntimeError(
            'No selected/relevant indentation measurements found'
        )

    curves = server.curves(doc_id=doc_id)
    analyses_classes = {
        item['class_id']: item
        for item in server.analyses.classes().get('result', [])
    }
    curve_type, x_dim, y_dim = choose_curve(curves)

    info(
        'Exporting selected indentation measurements from %s'
        % doc.get('name', doc_id)
    )
    info(' - curve type: %s' % curve_type)

    wb = Workbook(write_only=False)
    charts_ws = wb.active
    charts_ws.title = 'Charts'
    group_colors = {}
    summary_chart_data = write_results_sheet(
        wb,
        server,
        doc_id,
        selected,
        analyses_classes,
    )
    exported, x_name, x_unit, y_name, y_unit, curve_chart = write_curves_sheet(
        wb,
        server,
        doc_id,
        selected,
        curves,
        curve_type,
        x_dim,
        y_dim,
        group_colors,
    )
    average_curve_chart = write_average_curves_sheet(
        wb,
        exported,
        x_name,
        x_unit,
        y_name,
        y_unit,
        group_colors,
    )
    write_charts_sheet(
        charts_ws,
        os.path.basename(doc_path),
        summary_chart_data,
        curve_chart,
        average_curve_chart)

    wb.save(export_path)
    info('File saved: %s' % export_path)
    return export_path


def open_file_with_default_program(filename):
    if sys.platform.startswith('win'):
        os.startfile(filename)
    elif sys.platform == 'darwin':
        subprocess.Popen(['open', filename])
    else:
        subprocess.Popen(['xdg-open', filename])


def ask_to_open_file(filename):
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    should_open = messagebox.askyesno(
        'Open generated file?',
        'The Excel file was created successfully.\n\nOpen it now?',
        parent=root)
    root.destroy()

    if should_open:
        open_file_with_default_program(filename)


def check_for_updates():
    info('%s - Current version: %s' % (__title__, __version__))
    try:
        repo_path = __url__.rstrip('/').replace(
            'https://github.com/', '')
        api_url = (
            'https://api.github.com/repos/%s/tags' % repo_path)
        req = urllib.request.Request(
            api_url,
            headers={'User-Agent': 'opencode/1.0'},
            method='GET')
        with urllib.request.urlopen(req, timeout=10) as resp:
            tags = json.loads(resp.read().decode())
    except Exception:
        info(
            'Could not check for updates: no internet connection\n'
        )
        return

    latest = ''
    for tag in tags:
        ver = tag['name'].lstrip('v')
        if _version_greater(ver, latest):
            latest = ver

    if not latest:
        info('Could not determine latest version\n')
        return

    try:
        [int(x) for x in __version__.split('.')]
        current_valid = True
    except (ValueError, AttributeError):
        current_valid = False

    if not current_valid or _version_greater(__version__, latest):
        info('Could not determine latest version\n')
        return

    if _version_greater(latest, __version__):
        info(
            'New version %s available. Visit %s to download\n'
            % (latest, __url__)
        )
    else:
        info('Latest version (%s) is installed\n' % __version__)


def _version_greater(a, b):
    if not a:
        return False
    if not b:
        return True
    try:
        parts_a = [int(x) for x in a.split('.')]
        parts_b = [int(x) for x in b.split('.')]
    except (ValueError, AttributeError):
        return False
    max_len = max(len(parts_a), len(parts_b))
    parts_a += [0] * (max_len - len(parts_a))
    parts_b += [0] * (max_len - len(parts_b))
    for pa, pb in zip(parts_a, parts_b):
        if pa != pb:
            return pa > pb
    return False


if __name__ == '__main__':
    check_for_updates()
    indent = connect_to_indentation()
    result = indent.ls()
    info('Connected to %(server_name)s, V%(server_version)s' % result)

    docs = indent.docs()
    doc_id = docs.get('current') or docs['indexes'][0]
    filename = export_selected_indentation_excel(indent, doc_id)
    ask_to_open_file(filename)
